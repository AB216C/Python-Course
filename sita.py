print('####'*2)
print('#'*5)
print('#'*2)

patient_name = "john Smith"
age = 20
new_patient=True

name = input("What's your name?")
color = input("What's your favorite color?")
print(name + " likes " + color)

birth_year = input("what's your birth year?")
age = 2024-int(birth_year)
print(age)

weight = input("what is your weight(in pounds)?\n")
conv_kilogram = int(weight) * 0.453592
print(f"Your weight is:{conv_kilogram} kg")


print('I want Cakula cya "Uzima"')
print("I want Cakula cya 'Uzima'")

print('''
Hello John,

Your mail has been delivered.
Make sure you arrive at the front desk on time to pick it up.

Thank you!

Front desk Team.
''')

print("""
Hello Joseph,

Welcome to our team. Our goal is to do better that we used to.

Quality Team member
""")

course = "This course is called Python"
bishop = {"Car":"Ford", "Color":"Blue", "location":"South"}
item=(2,3,4,4,4,5,6)

print(course)
print(bishop)
print(item)

print(course[3])
print(course[0:20])
print(item[0:5])
print(item[:-1])
print(item[0:])
print(item[-1])
print(item[1:-1])
print(course[0:-1])

first = "John"
last = "Smith"
print(first + " [" + last + "]" + " is a coder")
print(f"{first} [{last}] is a coder")

quote = "Dream big, fight hard and prioritize integrity"
print(quote.upper())
print(quote.lower())
print(quote.find("fight"))
print(quote[11])
print(quote.replace("integrity","emotional intelligence"))
print(len(quote))
print(quote.title())
job_title=" quality assurance supervisor"
print(job_title.title())
print( 'big' in quote)

X = 10
Y = 10
X = X - 3
Y -= 3
print(X)
print(Y)

import math

print(math.sin(90))
print(math.cos(90))

x = 45
print(math.tan(x))

z = 2.67
y = 3.45
print(math.ceil(z))
print(math.floor(y))

#if statement using Comparison operators
temp = int(input("what is the temperature today(Degree celsius)? "))

if temp >= 27:
    print("It's a hot day")
    print("Drink plenty of water")
elif temp <= 10:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")

#if statement using boolean variable

hot = False
cold = False

if hot:
    print("It's a hot day")
    print("Drink plenty of water")
elif cold:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")

price = 1000000
good_credit = False

if good_credit:
    downpayment= price * 0.1
else:
    downpayment= price* 0.20

print(f"Your downpayment will be {downpayment}")


#if statement using logical operators(and, or, not)

good_credit = True
high_income = True

if good_credit and high_income:
    print("Eligible for a loan")
elif good_credit or high_income:
    print("You need a cosiner")
else:
    print("Not eligible for a loan")

good_credit = False
high_income = True
has_criminal_record=True

if good_credit and high_income:
    print("Eligible for a loan")
elif not has_criminal_record:
    print("eligible for a loan")
else:
    print("Not eligible for a loan")

#exercise on if statement with comparison operator

name = input("what'is your name?\n")
length_char=len(name)

if length_char <= 3:
    print("The name must be at least 3 characters")
elif length_char >= 50:
    print("The maximum characters is 50")
else:
    print("The name looks good. Please proceed to the next")

#weight conversion among different units

weight = int(input("What is your weight?\n"))
unit = input("what's the units of measurement K(Kg) or L(lbs)?\n")

if unit.upper() == "K":
    Total_weight = weight * 2.20462
else:
    Total_weight = weight * 0.453592

print(f"Your weight is: {Total_weight}")

#while loop. It works pretty much as if function. However, lines of code after while condition are  able to be repeated.
name = input("What's is your name?")

while name == "":
    print("Your name is not valid")
    name = input("Try again, What's is your name?")
print(f"Hello {name}! Good to see you.")


age = int(input("what's your age? \n"))

while age < 18:
    print("You are still young")
    age = int(input("Next. what's your age?\n"))

print(f"Your are {age} years old. Please enter inside")

#using logical operator not in while loop

brand = input("what's your car brand?\n")

while not brand == "suzuki":
    print("You got a new car")
    brand = input("Try again. what's your car brand?\n")
print("You got an old car")

#while loop. It works pretty much as if function. However, lines of code after while condition are  able to be repeated.
name = input("What's is your name?")

while name == "":
    print("Your name is not valid")
    name = input("Try again, What's is your name?")
print(f"Hello {name}! Good to see you.")


age = int(input("what's your age? \n"))

while age < 18:
    print("You are still young")
    age = int(input("Next. what's your age?\n"))

print(f"Your are {age} years old. Please enter inside")

#using logical operator not in while loop

brand = input("what's your car brand?\n")

while not brand == "suzuki":
    print("You got a new car")
    brand = input("Try again. what's your car brand?\n")
print("You got an old car")

# Building car game part 1 using while loops
command = ""
while command != "quit":
    command = input("> ").lower()
    if command == "start":
        print("cart started")
    elif command == "stop":
        print("car stopped")
    elif command == "help":
        print("""
    start - to start the car
    stop - to stop the car
    quit - to quit
        """)
    elif command == "quit":
        break
    else:
        print("sorry, i don't understand that")
#adding boolean to the program part 2
command = ""
while True:
    command = input("> ").lower()
    if command == "start":
        print("cart started")
    elif command == "stop":
        print("car stopped")
    elif command == "help":
        print("""
    start - to start the car
    stop - to stop the car
    quit - to quit
        """)
    elif command == "quit":
        break
    else:
        print("sorry, i don't understand that")

#adding extra information to the program: storing extra information in the memory part 3: Complete and ready to run
command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started")
        else:
            started = True
            print("cart started")
    elif command == "stop":
        if not started:
            print("car already stopped")
        else:
            started = False
            print("car stopped")
    elif command == "help":
        print("""
    start - to start the car
    stop - to stop the car
    quit - to quit
        """)
    elif command == "quit":
        break
    else:
        print("sorry, i don't understand that")

# practicing for loop

for item in "Python":
    print(item)

names = ["Jack", "Joseph", "Jane"]
for name in names:
    print(name)

numbers = range(10)
for number in numbers:
    print(number)
numbers = range(4,10)
for x in numbers:
    print(x)

for y in range(500, 1000,100):
    print(y)

# practicing for loop: Calculating SUM OF ELEMENTS IN A LIST
# This requires to include for loop elements

prices = [10,20, 30]

total = 0
for price in prices:
    total = total + price
print(total)

quantities = (30, 10, 15)

sum = 0
for quantity in quantities:
    sum = sum + quantity
print(sum)

# nested loops- loop in a loop

for x in range(4):
    for y in range(3):
        print(f"({x},{y})")

# challenge on for loops. Challenge 1: WRITING LETTER E from the list

numbers = [5, 2, 5, 2, 2]
i = 0
for i in numbers:
    print(i*("X"))
    i = i + 1
# Challenge on for loops. Challenge 2
items = [6, 2, 2, 6, 2, 2, 6]
x = 0
for x in items:
    print(x * "X")
    x = x + 1

#identifying the largest number in a list
#1st method:
my_list = [3,2,4,4,7,8,9]

print(my_list[0:])
print(my_list[:-1])

largest_number = max(my_list)
print(largest_number)
#2nd method-finding the max number using for loop
numbers = [4,5,6, 1, 10]

largest_n= numbers[0]

for number in numbers:
    if number > largest_n:
        largest_n = number

print(largest_n)
# find the maximum number using for loop: Comparing initial value with all other values and reset the value
list = [8, 90, 70, 56]

max = list[0]

for item in list:
    if item > max:
        max = item
print(max)


# 2-D List = Matrix and accessing elements in matrix
matrix = [
[1,2,3],
[2,5,0],
[5,3,7]
]
print(matrix[0][1])
print(matrix[0][2])
print(matrix[0][0])

for row in matrix:
    for item in row:
        print(item)

# list methods

numbers = [3, 4 , 5, 3, 8, 9]

print(8 in numbers)

numbers.pop()
numbers.insert(1,30)
numbers.append(70)
print(numbers)

numbers.sort()

print(numbers)



# removing duplicates from a list
# using "set" to remove duplicates

my_list = [1,2,2,4, 3, 5, 5, 9]

unique_list = list(set(my_list))

print(unique_list)

#removing duplicates using a loop and a temporary list

list_a = [3,4,5,6,6,8,9,7,7]

unique_list = []

for item in list_a:
    if item not in unique_list:
        unique_list.append(item)
print(unique_list)

#unpacking list and tuples

list = [1,3,6]
a,b,c = list
print(a)
print(b)
print(c)
tuple=(5,7,9)
x,y,z = tuple
print(x)
print(y)
print(z)

#Dictionaries

#using get method to call items in dictionaries

my_dict = {
    "Name": "Joseph Matata",
    "Age" : "34",
    "Birth_place":"Nyamasheke"
}

print(my_dict.get("Name"))
print(my_dict.get("Age"))

my_dict["birth_date"] = "June 1, 2000"
print(my_dict)

# challenge: convert numbers into characters such as one, two, three using python dictionary
#adding double quotes in the end for spacing of the output

phone_id = input("what's your phone digits?\n")

digits_map = {
    "0" : "Zero",
    "1" : "One",
    "2" : "Two",
    "3" : "Three",
    "4" : "Four",
    "5" : "Five",
    "6" : "six",
    "7" : "Seven",
    "8" : "Eight",
    "9" : "Nine"
}

output = ""
for char in phone_id:
    output += digits_map.get(char) + " "
print(output)

#using get method to get elements from a dictionary
message = input("Type your message\n")

words = message.split(" ")

emojis ={
    ":)": "Happy",
    ":(": "sad",
    "**": "alright",
    "&&": "cool"
}
output = ""

for word in words:
    output += emojis.get(word,word) + " "
print(output)


#functions-defining a function
#the function is executed all the way from the bottom

def greet_user():
    print("You are special passengers we have today")
    print("Please, Welcomed on board")

print("Hello All")
greet_user()

#parameters: Additng information to the function

def greet_user(name):
    print(f"Hello {name}")
    print("Welcome on board")

greet_user("John")

def greet_user(first_name,last_name):
    print(f"Hello {first_name} {last_name}")
    print("Welcome on board")


greet_user("Ferdinand", "Niyo")
print("Finish")

#function with Keywords-arguments-improving readability

def greet_user(first_name,last_name):
    print(f"Hello {first_name} {last_name}")
    print("Welcome on board")


greet_user(first_name="Ferdinand",last_name= "Niyo")
print("Finish")

#using return function to calculate the results of a given function

def square(number):
    return number * number

print(square(5))

def calculator(x):
    return x**5 -23

print(calculator(9))

def division(x):
    return x/6

print(division(100))

#dealing with exceptions: To prevent the program form crashing

try:
    age = int(input("How old are you?\n"))
    print(age)
    income = 1000
    risk = income/age
    print(risk)
except ValueError:
    print("Invalid input")
except ZeroDivisionError:
    print("Invalid value")


    # # classes-Basically provide a room to store several information regarding certain item
    #
    class Person:
        def __init__(self, name):
            self.name = name

        def talk(self):
            print("name")
            print("my names and my heroes")

        def address(self):
            print("my email is @gmail.com")
            print("my phone is 33455")

        def course(self):
            print("Chemistry and Math")
            print("Geograph and Education Civique")


    john = Person("John Smith")
    print(john.name)
    john.talk()
    john.address()
    john.course()


    class Apartment:
        def __init__(self, room):
            self.room = room

        def kitchen(self):
            print("plates, spoons, cleaning tools, stoves, and utensils")

        def living_room(self):
            print("couches, tables, TV, Computer")

        def bed_room(self):
            print("King mattress, romantic lights, aestetic")


    roomA = Apartment("roomA1")
    print(roomA.room)
    roomA.kitchen()
    roomA.living_room()
    roomA.bed_room()


    # adding parameters to the class

    class Apartment:
        def __init__(self, room):
            self.room = room

        def kitchen(self):
            print(
                f"We prefer to have kitchen for {self.room} full of plates, spoons, cleaning tools, stoves, and utensils")

        def living_room(self):
            print("couches, tables, TV, Computer")

        def bed_room(self):
            print(f" For all {self.room} should have King mattress, romantic lights, aestetic")


    roomA = Apartment("roomA1")
    roomA.bed_room()
    roomA.kitchen()

    roomB = Apartment("roomB1")
    roomB.bed_room()
    roomB.kitchen()
import random

for i in range(3):
    print(random.random())

leaders = ["Mary","John", "Joseph"]

print(random.choice(leaders))


for i in range(5):
    print(random.randint(7,100))

#rolling a dice: printing a pair of numbers(x,y)

number_of_pair = 1

for i in range(number_of_pair):
    x = random.randint(0,9)
    y = random.randint(0,9)

print(f"({x},{y})")

#practicing classes
class Dog:
    def __init__(self,food):
        self.food = food
        print(f"Most of dog's {food} are not humanly edible")


    def nutrients(self):
        print(f"most dog's food nutrients comes form {self.food}")

    def diseases(self):
        print(f"dog can be cured with {self.food}")

dog1 = Dog("bones")
dog1.nutrients()
dog1.diseases()


#finding the maximum from a list
list = [4,6,7,4,100,400]

maximum = list[0]

for item in list:
    if item > maximum:
        maximum = item
print(maximum)

#Matching up 2 basketball teams ranked to the top in preliminary competition: Paris Olympics, teams ready for Quarter final: First knockout session
#MEN
import random
print("Men's basketball teams: Quarter final shuffle")
print("\n")

teams = ["Canada", "Serbia", "France", "Australia", "Greece", "Germany","Brazil", "United States"]
random.shuffle(teams)

pairs = [(teams[i], teams[i+1]) for i in range(0, len(teams),2)]
for pair in pairs:
    print(f"{pair[0]} vs {pair[1]}")

#WOMEN
print("\n")
print("Woman Basketball Teams: Quarter Final shuffle\n")

teams2 = ["United States","Nigeria","Spain","Serbia","France","Australia","Germany","Belgium"]
random.shuffle(teams2)
pairs = [(teams2[i],teams2[i+1]) for i in range(0,len(teams2),2)]

for pair in pairs:
    print(f"{pair[0]} vs {pair[1]}")

class Gataraga:
    def __init__(self,attractions):
        self.attractions = attractions
        print(f"I love Gataraga main beauties such as {attractions}")

    def market(self):
        print(f"{self.attractions} attract a lot of people whose markets ideas ")

    def imigezi(self):
        print(f"Dukunda umugezi wa mutobo. uri hafi ya {self.attractions}")

mountain = Gataraga("Mugombwa")
mountain.market()
mountain.imigezi()


#practicing Inheritance all the way from class module in python
class Birds():
    def kuguruka(self):
        print("Kiraguruka")
class Igishwi(Birds):
    def Kwabira(self):
        print("Igishwi, Kirabira")
class Ikinyamanza(Birds):
    pass


bird1 = Ikinyamanza()
bird1.kuguruka()

bird3=Igishwi()
bird3.Kwabira()

class Amphibians():
    def jump(self):
        print("All Amphibians are like that")

class Frog():
    def distance(self):
        print("Most frogs jumps in a long distance")

class Toad():
    def speed(self):
        print("Most toads jump slow enough to catch")

amph1 = Amphibians()
amph1.jump()

amph2 = Frog()
amph2.distance()

amp3 = Toad()
amp3.speed()


#Rolling dice exercise
import random


class Dice:

    def roll(self):
        first_number = random.randint(1,6)
        second_number = random.randint(1,6)
        return first_number,second_number

dice = Dice()
print(dice.roll())


#working with files using pathlib module
from pathlib import Path

path = Path()

 for file in path.glob('*.py'):
     print(file)

for file in path.glob('*'):
    print(file)

path = Path("ecommerce")
print(path.exists())

path = Path("emails")
print(path.exists())

#working with excel

#pypi.org: very important in installing differnt python functions
# to use excel: term

import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']

print(cell.value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    print(cell.value)
    corrected_price = cell.value * 0.9

